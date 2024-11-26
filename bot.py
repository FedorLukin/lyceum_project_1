import datetime as dt
import telebot
import openpyxl
import time
import os
from cachetools import TTLCache
from dotenv import load_dotenv
from typing import Union
from db.models import regular_schedule
from db.models import uday_schedule
from db.models import users
from telebot.apihelper import ApiTelegramException
from telebot import types
import logging


logs_format = '%(asctime)s - %(filename)s:%(lineno)d - %(message)s'
logging.basicConfig(level=logging.ERROR, filename='logs.log', filemode='w', format=logs_format)
load_dotenv()
bot = telebot.TeleBot(os.getenv('TELEGRAM_BOT_TOKEN_APIKEY'))
messages_cache = TTLCache(maxsize=400, ttl=5)
errors_cache = TTLCache(maxsize=10, ttl=100)


def caching_decorator(func: callable) -> callable:
    """
    Декоратор, кэширующий запросы пользователя к боту.

    Декоратор ограничивает количество запросов от пользователя,
    если запрос не был выполнен пользователем недвано выполняется оборачиваемая функция,
    иначе функция не выполняется.

    Аргументы:
        func (callable): Декорируемая функция.
    
    Возвращает:
        (callable): Декорированная функция, принимающая сообщение пользователя.
        """
    def wrapped(*args):
        message = args[0]
        user_id = message.from_user.id
        if not messages_cache.get(user_id):
            messages_cache[user_id] = True
            return func(message)
    return wrapped


def parent_of_merged_cell(cell: openpyxl.cell.cell.MergedCell) -> str:
    """
    Функция, ищущая родителя объединённой клетки таблицы.

    Аргументы:
        cell (openpyxl.cell.cell.MergedCell): Объединённая клетка.
    
    Возвращает:
        str: Координаты родительской клетки.
    """
    sheet = cell.parent
    child_coord = cell.coordinate
    for merged in sheet.merged_cells.ranges:
        if child_coord in merged:
            return merged.start_cell.coordinate


def cell_value(cell: Union[openpyxl.cell.cell.Cell, openpyxl.cell.cell.MergedCell]) -> str:
    """
    Функция возвращающая значение клетки.

    Функция проверяет тип клетки, возвращает её значение если это обычная клетка, ищет родительскую клетку
    и возвращает её значение если клетка совмещённая.
    
    Аргументы:
        cell Union[openpyxl.cell.cell.Cell, openpyxl.cell.cell.MergedCell]: Клетка с искомым значением.

    Возвращает:
        str: Значение клетки.
    """
    if isinstance(cell, openpyxl.cell.cell.Cell):
        return cell.value
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        coord = parent_of_merged_cell(cell)
        parent = cell.parent[coord]
        return parent.value


def regular_classes_schedule_parsing(date: dt.date, worksheet: openpyxl.worksheet.worksheet.Worksheet,
                                     times_list: list, start_row: int, start_col: int, end_row: int,
                                     end_col: int) -> None:
    """
    функция парсинга обычного расписания классов.

    Функция итерируется по столбцам и клеткам столбца, читает расписание и сохраняет его в базу данных.

    Аргументы:
        date (dt.date): Дата из файла с расписанием.
        worksheet (openpyxl.worksheet.worksheet.Worksheet): Страница эксель файла.
        times_list (list): Список с таймингами уроков.
        start_row (int): Начальная строка итерации.
        start_col (int): Начальный столбец итерации.
        end_row (int): Конечная строка итерации.
        end_col (int): Конечный столбец итерации.

    Возвращает:
        None: функция ничего не возвращает.
    """
    for col in worksheet.iter_cols(min_row=start_row, min_col=start_col, max_row=end_row, max_col=end_col):
        key = cell_value(col[0])
        groups = ('гр.А', 'гр.Б')
        group = groups.index(''.join(cell_value(col[1]).split()))
        for i, cell in enumerate(col[2:]):
            if cell_value(cell):
                lesson_info = '\n'.join((times_list[i], '\n'.join(cell_value(cell).split('\n\n'))))
                regular_schedule.objects.create(lesson_number=i, lesson_info=lesson_info, class_letter=key,
                                                group_number=group, date=date)


def uday_groups_schedule_parsing(date: dt.date, worksheet: openpyxl.worksheet.worksheet.Worksheet,
                                 times_list: list, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    """
    функция парсинга расписания для групп на универдень.

    Функция итерируется по столбцам и клеткам столбца, читает расписание и сохраняет его в базу данных.

    Аргументы:
        date (dt.date): Дата из файла с расписанием.
        worksheet (openpyxl.worksheet.worksheet.Worksheet): Страница эксель файла.
        times_list (list): Список с таймингами уроков.
        start_row (int): Начальная строка итерации.
        start_col (int): Начальный столбец итерации.
        end_row (int): Конечная строка итерации.
        end_col (int): Конечный столбец итерации.

    Возвращает:
        None: функция ничего не возвращает.
    """
    done = set()
    for col in worksheet.iter_cols(min_row=start_row, min_col=start_col, max_row=end_row, max_col=end_col):
        group = int(cell_value(col[0]).split()[0])
        if group in done:
            continue
        done.add(group)
        for i, cell in enumerate(col[1:]):
            if cell_value(cell):
                lesson_info = '\n'.join((times_list[i], '\n'.join(cell_value(cell).split('\n\n'))))
                uday_schedule.objects.create(lesson_number=i, lesson_info=lesson_info, group_number=group, date=date)


def uday_classes_schedule_parsing(date: dt.date, worksheet: openpyxl.worksheet.worksheet.Worksheet,
                                  times_list: list, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    """
    функция парсинга расписания для классов на универдень.

    Функция итерируется по столбцам и клеткам столбца, читает расписание и сохраняет его в базу данных.

    Аргументы:
        date (dt.date): Дата из файла с расписанием.
        worksheet (openpyxl.worksheet.worksheet.Worksheet): Страница эксель файла.
        times_list (list): Список с таймингами уроков.
        start_row (int): Начальная строка итерации.
        start_col (int): Начальный столбец итерации.
        end_row (int): Конечная строка итерации.
        end_col (int): Конечный столбец итерации.

    Возвращает:
        None: функция ничего не возвращает.
    """
    done = set()
    for col in worksheet.iter_cols(min_row=start_row, min_col=start_col, max_row=end_row, max_col=end_col):
        key = cell_value(col[0])
        if key in done:
            continue
        done.add(key)
        for i, cell in enumerate(col[1:]):
            if cell_value(cell):
                lesson_info = '\n'.join((times_list[i], '\n'.join(cell_value(cell).split('\n\n'))))
                regular_schedule.objects.create(lesson_number=i, lesson_info=lesson_info, class_letter=key,
                                                group_number=0, date=date)


def main_schedule_parse(filename: str) -> str:
    """
    Основная функция парсинга расписания.

    Функция открывает файл с расписанием и вызывает вспомогательные функции для парсинга расписания в зависимости
    от дня недели. В случае успешного парсинга функция рассылает пользователям уведомление о загрузке расписания.

    Аргументы:
        filename (str): Имя открываемого файла.
    
    Возвращает:
        str: Сообщение об успехе или ошибке в ходе выполнения функции.

    """
    workbook = openpyxl.load_workbook(f'./uploads/{filename}')
    os.remove(f'./uploads/{filename}')
    today = dt.date.today()
    old_date = today - dt.timedelta(days=2)
    date = dt.datetime.strptime(f'{filename.split('.xlsx')[0]}{today.year}', "%d.%m%Y").date()
    weekday = dt.date.weekday(date)
    regular_schedule.objects.filter(date__lt=old_date).delete()
    uday_schedule.objects.filter(date__lt=old_date).delete()
    sh_10, sh_11 = None, None
    for i, sh in enumerate(workbook.sheetnames):
        if sh.strip() == '10':
            sh_10 = i
        elif sh.strip() == '11':
            sh_11 = i
        if sh_10 and sh_11:
            break

    #  удаление записей при повторной загрузке расписания
    if regular_schedule.objects.filter(date=date).exists():
        regular_schedule.objects.filter(date=date).delete()
        uday_schedule.objects.filter(date=date).delete()

    # париснг расписания 10-классников
    try:
        # универ-день
        if weekday == 0:
            sheet = workbook.worksheets[0] if not sh_10 else workbook.worksheets[sh_10]
            times_10 = [cell_value(i).replace('\n', '') for i in sheet['c'][2:9] + sheet['c'][9:11] if cell_value(i)]
            uday_groups_schedule_parsing(date, sheet, times_10[:6], 2, 4, 8, 27)
            uday_classes_schedule_parsing(date, sheet, times_10[6:], 9, 4, 12, 27)

        # все остальные дни недели
        else:
            sheet = workbook.worksheets[1] if not isinstance(sh_10, int) else workbook.worksheets[sh_10]
            ls_num = max([int(cell_value(i)) for i in sheet['B'][3:14] if str(cell_value(i)) in '123456789'])
            times_10 = [cell_value(i).replace('\n', '') for i in sheet['c'][3:3 + ls_num]]
            regular_classes_schedule_parsing(date, sheet, times_10, 2, 4, 11, 23)
    except Exception as ex:
        regular_schedule.objects.filter(date=date).delete()
        uday_schedule.objects.filter(date=date).delete()
        logging.error(ex)
        return f'Ошибка при парсинге расписания 10-х классов!\nОшибка:\n{ex}'

    # парсинг расписания 11-классников:
    # универ-день
    try:
        if weekday == 2:
            sheet = workbook.worksheets[0] if not sh_11 else workbook.worksheets[sh_11]
            times_11 = [cell_value(i).replace('\n', '') for i in sheet['c'][3:9] + sheet['c'][10:12] if cell_value(i)]
            uday_groups_schedule_parsing(date, sheet, times_11[:6], 3, 4, 9, 23)
            uday_classes_schedule_parsing(date, sheet, times_11[6:], 10, 4, 13, 23)

        # все остальные дни недели
        else:
            sheet = workbook.worksheets[1] if not isinstance(sh_11, int) else workbook.worksheets[sh_11]
            times_11 = [cell_value(i).replace('\n', '') for i in sheet['c'][3:11] if cell_value(i)]
            regular_classes_schedule_parsing(date, sheet, times_11, 2, 4, 12, 23)
    except Exception as ex:
        regular_schedule.objects.filter(date=date).delete()
        uday_schedule.objects.filter(date=date).delete()
        logging.error(ex)
        return f'Ошибка при парсинге расписания 11-х классов!\nОшибка:\n{ex}'

    # рассылка уведомления о загрузке расписания
    for user in users.objects.all():
        kb = types.InlineKeyboardMarkup()
        kb.add(types.InlineKeyboardButton('расписание на сегодня', callback_data='get_schedule=today'))
        kb.add(types.InlineKeyboardButton('расписание на завтра', callback_data='get_schedule=tommorow'))
        try:
            bot.send_message(user.user_id, f'Загружено расписание на {date}', reply_markup=kb)
        except ApiTelegramException as ex:
            if ex.description == 'Forbidden: bot was blocked by the user':
                user.delete()
        time.sleep(0.036)
    return 'Расписание сохранено успешно!'


def confirm_notification(message: telebot.types.Message, recievers: str) -> None:
    """
    Функция подтверждения содержимого рассылаемого сообщения.
        
    Функция запрашивает у админа подтверждение запуска рассылки и в зависимости от ответа
    запускает или отменяет рассылки.

    Аргументы:
        message (telebot.types.Message): Сообщение отправленное ранее админом.
        recievers (str): Получатели сообщения.

    Возвращает:
        None: Функция ничего не возвращает.
    """
    kb = types.InlineKeyboardMarkup(row_width=1)
    kb.add(types.InlineKeyboardButton('отправить', callback_data=f'send={recievers}'),
           types.InlineKeyboardButton('отмена', callback_data='back_to_admin'))
    if message.photo:
        photo_bytes = bot.download_file(bot.get_file(message.photo[-1].file_id).file_path)
        text = message.caption if message.caption else ''
        bot.send_photo(message.from_user.id, photo=photo_bytes,
                       caption=f'подвердите отправку сообщения\n>получатель:\n{recievers}\n>сообщение:\n{text}',
                       reply_markup=kb)
    else:
        bot.send_message(message.from_user.id,
                         f'подвердите отправку сообщения\n>получатель:\n{recievers}\n>сообщение:\n{message.text}',
                         reply_markup=kb)


def schedule_adding(message: telebot.types.Message) -> None:
    """
    Функция получения файла с расписанием.
    
    Функция получает сообщение от админа, если оно содержит .xlsx файл - вызывает функцию парсинга, иначе отправляет
    сообщение об ошибке.

    Аргументы:
        message (telebot.types.Message): Сообщение с файлом расписания отправленное админом.
    
    Возвращает:
        None: Функция ничего не возвращает.
    """
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton('вернуться к админ-панели', callback_data='back_to_admin'))
    if message.content_type == 'document' and message.document.file_name.endswith('.xlsx'):
        file_name = message.document.file_name
        file_info = bot.get_file(message.document.file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        with open(f'uploads/{file_name}', 'wb') as new_file:
            new_file.write(downloaded_file)
        bot.send_message(message.from_user.id, main_schedule_parse(file_name), reply_markup=kb)
    else:
        kb.add(types.InlineKeyboardButton('попробовать ещё раз', callback_data='add_schedule'))
        bot.send_message(message.from_user.id, 'Файл с расписанием должен быть формата .xlsx, попробуйте снова',
                         reply_markup=kb)


@bot.message_handler(commands=['get'])
@caching_decorator
def get(message: telebot.types.Message) -> None:
    """
    Функция ответа на запрос расписания.

    Функция отправляет пользователю в ответ сообщение, запрашивая день на которой необходимо отправить расписание.

    Аргументы:
        message (telebot.types.Message): Сообщение отправленное пользователем.

    Возвращает:
        None: Функция ничего не возвращает.
    """
    if users.objects.filter(user_id=message.from_user.id).exists():
        kb = types.InlineKeyboardMarkup()
        kb.add(types.InlineKeyboardButton('расписание на сегодня', callback_data='get_schedule=today'))
        kb.add(types.InlineKeyboardButton('расписание на завтра', callback_data='get_schedule=tommorow'))
        bot.send_message(message.from_user.id, 'выберите действие', reply_markup=kb)


@bot.message_handler(commands=['start', 'edit'])
@caching_decorator
def start(message: telebot.types.Message) -> None:
    """
    Функция обработки команд /start и /edit.

    Функция отправляет пользователю в ответ сообщение с предложением заполнить данные.

    Аргументы:
        message (telebot.types.Message): Сообщение отправленное пользователем.
    
    Возвращает:
        None: Функция ничего не возвращает.
    """
    command = message.text
    if command == '/start' and not users.objects.filter(user_id=message.from_user.id).exists() or command == '/edit':
        kb = types.InlineKeyboardMarkup()
        kb.add(types.InlineKeyboardButton('начать', callback_data='choice'))
        bot.send_message(message.from_user.id, 'Привет, давай определимся с твоими классом и группой', reply_markup=kb)


@bot.message_handler(commands=['admin'])
def admin_panel(message: telebot.types.Message) -> None:
    """
    Функция админ-панели.

    Функция проверяет является ли пользователь админом бота и предоставляет доступр к админ-панели если это так,
    иначе запрос пользователя игнорируется.

    Аргументы:
        message (telebot.types.Message): Сообщение отправленное поьзователем.
    
    Возвращает:
        None: Функция ничего не возвращает.
    """
    if str(message.from_user.id) == os.getenv('ADMIN_ID'):
        kb = types.InlineKeyboardMarkup(row_width=1)
        kb.add(types.InlineKeyboardButton('Добавить расписание', callback_data='add_schedule'),
               types.InlineKeyboardButton('Сделать рассылку', callback_data='make_notification'))
        bot.send_message(message.from_user.id, 'Добро пожаловать в админ-панель! Выберите действие на клавиатуре',
                         reply_markup=kb)


@bot.callback_query_handler(func=lambda callback: True)
def callback_message(callback: telebot.types.CallbackQuery) -> None:
    """
    Функция обработки пользовательских нажатий на кнопки.

    Функция отвечает на запросы пользователя, сделанные посредством нажатий кнопок.

    Аргументы:
        callback (telebot.types.CallbackQuery): Коллбэк вызванный нажатием на кнопку.
    
    Возвращает:
        None: Функция ничего не возвращает.
    """
    # выбор цифры класса
    if callback.data == 'choice':
        keyboard = types.InlineKeyboardMarkup(row_width=2)
        keyboard.add(types.InlineKeyboardButton('10', callback_data='10'),
                     types.InlineKeyboardButton('11', callback_data='11'))
        bot.edit_message_text('выберите класс', callback.from_user.id, callback.message.message_id,
                              reply_markup=keyboard)

    # выбор буквы класса
    elif callback.data == '10' or callback.data == '11':
        cl_num = callback.data
        classes_11 = [('В(бета)', 'Η(эта)'), ('Ζ(дзeта)', 'Θ(тета)'), ('Г(гамма)', 'Ε(эпсилон)'),
                      ('Ι(йота)', 'К(каппа)'), ('Δ(дельта)', 'Λ(лямбда)')]
        classes_10 = [('Μ(мю)', 'Σ(сигма)'), ('Ξ(кси)', 'Τ(тау)'), ('Ο(омикрон)', 'Φ(фи)'), ('Π(пи)', 'Х(хи)'),
                      ('Ρ(ро)', 'Ψ(пси)')]
        classes = classes_10 if cl_num == '10' else classes_11
        keyboard = types.InlineKeyboardMarkup(row_width=2)
        for cl1, cl2 in classes:
            keyboard.add(types.InlineKeyboardButton(cl1, callback_data=f'class_letter={cl_num} {cl1.split('(')[0]}'),
                         types.InlineKeyboardButton(cl2, callback_data=f'class_letter={cl_num} {cl2.split('(')[0]}'))
        bot.edit_message_text('выберите букву', callback.from_user.id, callback.message.message_id,
                              reply_markup=keyboard)

    # выбор группы класса
    elif callback.data.startswith('class_letter'):
        cl_letter = callback.data.split('=')[1]
        user = users.objects.get_or_create(user_id=callback.from_user.id, defaults={'class_letter': ''})[0]
        user.class_letter = cl_letter
        user.save(update_fields=['class_letter'])
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(types.InlineKeyboardButton('группа А', callback_data='class_group=группа А'),
                     types.InlineKeyboardButton('группа Б', callback_data='class_group=группа Б'))
        bot.edit_message_text('выберите группу', callback.from_user.id, callback.message.message_id,
                              reply_markup=keyboard)

    # выбор группы на универ-день
    elif callback.data.startswith('class_group'):
        cl_group = callback.data.split('=')[1]
        cl_group = 0 if cl_group == 'группа А' else 1
        user = users.objects.get(user_id=callback.from_user.id)
        user.group_number = cl_group
        user.save(update_fields=['group_number'])
        cl = user.class_letter.split()[0]
        i, j = (6, 5) if cl == '11' else (7, 6)
        keyboard = types.InlineKeyboardMarkup(row_width=2)
        for i in range(1, i):
            keyboard.add(types.InlineKeyboardButton(str(i), callback_data=f'univer_group={str(i)}'),
                         types.InlineKeyboardButton(str(i + j), callback_data=f'univer_group={str(i + j)}'))
        bot.edit_message_text('выберите группу универдня', callback.from_user.id, callback.message.message_id,
                              reply_markup=keyboard)

    # подтверждение данных
    elif callback.data.startswith('univer_group'):
        univer_group = int(callback.data.split('=')[1])
        user = users.objects.get(user_id=callback.from_user.id)
        user.u_group_number = univer_group
        user.save(update_fields=['u_group_number'])
        cl_letter, cl_group = user.class_letter, ['Гр. А', 'Гр. Б'][user.group_number]
        kb = types.InlineKeyboardMarkup(row_width=1)
        kb.add(types.InlineKeyboardButton('заполнить заново', callback_data='choice'),
               types.InlineKeyboardButton('сохранить', callback_data='done'))
        bot.edit_message_text(f'вы выбрали:\n{cl_letter} класс\n{cl_group}\n{univer_group} группа универдня',
                              callback.from_user.id, callback.message.message_id, reply_markup=kb)

    # уведомление об успешном сохранении записи
    elif callback.data == 'done':
        kb = types.ReplyKeyboardMarkup()
        kb.row('/edit', '/get')
        bot.edit_message_text('Успешно сохранено!\n/edit - заполнить заново\n/get - получить расписание',
                              callback.from_user.id, callback.message.message_id)

    # получение файла с расписанием
    elif callback.data == 'add_schedule':
        kb = types.InlineKeyboardMarkup()
        kb.add(types.InlineKeyboardButton('назад', callback_data='back_to_admin'))
        bot.edit_message_text('Отправьте файл чтобы добавить расписание', callback.from_user.id,
                              callback.message.message_id, reply_markup=kb)
        bot.register_next_step_handler(callback.message, schedule_adding)

    # возврат к админ-панели
    elif callback.data == 'back_to_admin':
        bot.clear_step_handler_by_chat_id(callback.message.chat.id)
        kb = types.InlineKeyboardMarkup(row_width=1)
        kb.add(types.InlineKeyboardButton('Добавить расписание', callback_data='add_schedule'),
               types.InlineKeyboardButton('Сделать рассылку', callback_data='make_notification'))
        if callback.message.photo:
            bot.delete_message(callback.from_user.id, callback.message.id)
            bot.send_message(callback.from_user.id,
                             'Добро пожаловать в админ-панель! Выберите действие на клавиатуре', reply_markup=kb)
        else:
            bot.edit_message_text('Добро пожаловать в админ-панель! Выберите действие на клавиатуре',
                                  callback.from_user.id, callback.message.message_id, reply_markup=kb)

    # выбор адресатов рассылки
    elif callback.data == 'make_notification':
        kb = types.InlineKeyboardMarkup(row_width=2)
        kb.add(types.InlineKeyboardButton('10е классы', callback_data='ntf=10'),
               types.InlineKeyboardButton('11е классы', callback_data='ntf=11'))
        kb.add(types.InlineKeyboardButton('отправить всем', callback_data='ntf=all'))
        kb.add(types.InlineKeyboardButton('назад', callback_data='back_to_admin'))
        bot.edit_message_text('выберите получателя', callback.from_user.id, callback.message.message_id,
                              reply_markup=kb)

    # запрос сообщения для рассылки
    elif callback.data.startswith('ntf'):
        recievers = callback.data.split('=')[1]
        kb = types.InlineKeyboardMarkup()
        kb.add(types.InlineKeyboardButton('назад', callback_data='back_to_admin'))
        bot.edit_message_text('Отправьте сообщение чтобы сделать рассылку', callback.from_user.id,
                              callback.message.message_id, reply_markup=kb)
        bot.register_next_step_handler(callback.message, confirm_notification, recievers)

    # рассылка сообщения
    elif callback.data.startswith('send'):
        bot.delete_message(callback.from_user.id, callback.message.message_id)
        recievers = users.objects.all() if callback.data.split('=')[1] == 'all' else users.objects.filter(
            class_letter__startswith=callback.data.split('=')[1])
        message = callback.message
        text = message.caption.split('>сообщение:')[1] if message.photo and message.caption else ''
        photo_bytes = bot.download_file(bot.get_file(message.photo[-1].file_id).file_path) if message.photo else None
        for user in recievers:
            try:
                if photo_bytes:
                    bot.send_photo(user.user_id, caption=text, photo=photo_bytes)
                else:
                    bot.send_message(user.user_id, message.text.split('>сообщение:')[1])
            except ApiTelegramException as ex:
                if ex.description == 'Forbidden: bot was blocked by the user':
                    user.delete()
            time.sleep(0.036)

    # отправка расписания
    elif callback.data.startswith('get_schedule'):
        day = callback.data.split('=')[1]
        kb = types.InlineKeyboardMarkup()
        kb.add(types.InlineKeyboardButton('расписание на сегодня', callback_data='get_schedule=today'))
        kb.add(types.InlineKeyboardButton('расписание на завтра', callback_data='get_schedule=tommorow'))
        date = dt.date.today() if day == 'today' else dt.date.today() + dt.timedelta(days=1)
        user = users.objects.get(user_id=callback.from_user.id)
        uday_flag = True if user.class_letter.startswith('11') and date.weekday() == 2 or user.class_letter.startswith(
            '10') and date.weekday() == 0 else False
        gr_num = 0 if uday_flag else user.group_number
        schedule_list = []
        if uday_schedule.objects.filter(date=date).exists() and uday_flag:
            schedule_list.append('\n\n'.join(
                [row.lesson_info for row in uday_schedule.objects.filter(group_number=user.u_group_number, date=date)]))
        if regular_schedule.objects.filter(date=date).exists():
            schedule_list.append('\n\n'.join([row.lesson_info for row in
                                              regular_schedule.objects.filter(class_letter=user.class_letter,
                                                                              group_number=gr_num, date=date)]))
        if schedule_list:
            kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
            kb.row('/edit', '/get')
            bot.send_message(callback.from_user.id, '\n\n'.join(schedule_list), reply_markup=kb)
            bot.delete_message(callback.from_user.id, callback.message.id)
        elif callback.message.text == 'выберите действие' or date.strftime("%d.%m") != \
                callback.message.text.split('расписание на ')[1][:5]:
            bot.edit_message_text(f'расписание на {date.strftime("%d.%m")} ещё не добавлено\nвыберите действие',
                                  callback.from_user.id, callback.message.message_id, reply_markup=kb)


if __name__ == '__main__':
    while True:
        try:
            bot.polling(none_stop=True)
        except Exception as ex:
            error_name = ex.__class__.__name__
            if not errors_cache.get(error_name):
                errors_cache[error_name] = True
                logging.error(ex)
